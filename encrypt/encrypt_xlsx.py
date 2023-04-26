import openpyxl
import os
from Cryptodome.Cipher import AES
import base64

# xlsx文件路径
XLSX_PATH = './encrypt/attach.xlsx'
# 输出的xlsx文件路径
OUTPUT_PATH = './encrypt/output.xlsx'
# 密钥
KEY = 'abc'
# 渠道列
ID_COL = 1
# 渠道名称列
CHANNEL_COL = 2
# attach列
ATTACH_COL = 11

# aes加密
def encrypt(str, key):
    if not key:
        raise ValueError("key不能为空")
    try:
        if not str:
            return None
        # 判断Key是否为16位
        if len(key) != 16:
            return None
        raw = key.encode('utf-8')
        skeySpec = AES.new(raw, AES.MODE_ECB)
        # "算法/模式/补码方式"
        # PKCS#7 padding
        BS = AES.block_size
        padded_str = str.encode('utf-8') + (BS - len(str) % BS) * chr(BS - len(str) % BS).encode('utf-8')
        encrypted = skeySpec.encrypt(padded_str)
        # 此处使用BASE64做转码功能，同时能起到2次加密的作用。
        return base64.b64encode(encrypted).decode('utf-8')
    except Exception as ex:
        return None

# 解密
def decrypt(str, key):
    if not key:
        raise ValueError("key不能为空")
    try:
        if not str:
            return None
        # 判断Key是否为16位
        if len(key) != 16:
            return None
        raw = key.encode('utf-8')
        cipher = AES.new(raw, AES.MODE_ECB)
        # 先用base64解密
        encrypted = base64.b64decode(str)
        try:
            original = cipher.decrypt(encrypted)
            originalString = original.decode('utf-8')
            return originalString
        except Exception as e:
            print("decrypt error", e)
            return None
    except Exception as ex:
        print(ex)
        return None
    
# 加密attach
def encrypt_attach(file_name, ouput_path):
    # 打开xlsx文件
    wb = openpyxl.load_workbook(file_name)
    # 获取第一个sheet
    sheet = wb.active
    rows = []
    # 遍历每一行数据
    for row in sheet.iter_rows():
        print(f"row length: {len(row)}")
        # 加密每行数据的attach列
        channel = row[ID_COL-1].value
        channel_name = row[CHANNEL_COL-1].value
        attach = row[ATTACH_COL-1].value
        print(channel, channel_name, attach)
        rows.append([channel, channel_name, attach, encrypt(attach, KEY)])
    # 写入到xlsx文件中
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in rows:
        sheet.append(row)
    wb.save(ouput_path)

if __name__ == '__main__':
    # 当前路径
    print(os.getcwd())

    # 读xlsx文件
    # encrypt_attach(XLSX_PATH, OUTPUT_PATH)

    # 解码
    print(decrypt('xZdT9O6k294n4HIFeowbNWbngs6UKCwhN9y9j2DDJqZ7k2oWgnGddNH1flg04YkFoDqBlFP4b/ojn1qTnEmHlFO2tMGL57y8ZSUuax7vbteHzsmOOZYLE36+oRu0BB3kCcxl9luOCEnVZ4NccLMh4S56svg1xBBWhm54qvor/OM3VLOd1JIkv8zldMvSwCdSvK24xoZ3B0WWOvN4lTHkQ0j6JJE1Iq0kDl9TKjIYSAX+44mf0DCFacGXRUxUUHR7E/366dtHA3Cow5D388wupuomCCwMBTwYrhcrj2cbZDLDsWwN0G6QxBIHbl62/15M+08YiLlcCZKUkOLvQDYt6FeWjHkDpTvB0nDSBkTSC9isfCw877oa6yhzuPFPxe7picKibBWexMyXUEslwFktTtYb5QvKScVgSv57O5So7fHYOkelswHBX3U5qjA7r2AEvLmefbGbPJILBoEm37astkpCMcjunDZy/Ir6aG5JfecYf9DtNGDRQBUmU9ls66myPUopNpW5dJbxcIyu1Ob62Da4g1KdOQfmo84ChGzHAJZlpgoa0B6zIBQpqqhIHg9LQEulcVbc4AOk2yc9pHb1LqqZiTV99IkdwFoJeK5KA8/333OULPQYi/n2xhXmX66PQUOCJx6aRyBvt1GdYdHA+EKkV3fl6BYuT53JrdOluzgsQ19D8OaWJ4+lHqwm+s/sG08d6awYhPgZ+WaC5tE6PFefY8x5Zgbi0Im2eawvZ1Z0/G8uNPhqesK5jP8vZ0srool8J6ITw+QqVH5ks7dypgS3HSusSeLPhL84I3Xik9i/pwR9IKSJtHQ1DSt8DwUD/hNGS1GCPOjbUg+u0a2EPHRA546Fee68mk4MYtWE6QyMQp1DRTepgRk3sbKoc+09j09qnQTcZukMXOoHCpUme0X2hOpuMgiTd+fgekzH2+OVxMG1WWa1X2jEJOB6EN4uIDDGBVb86f7JiH1VyDArVEpvJ5+d8ItGuE5k0J8pktXP4tBwV2A+21835+7xHn3NrxUgUtOoBr9gAmEP38XRSf2qRG6qNgFKuPhR7AqRkL9JzjV48kpOqvfMcKeew0tnWmZAiNmUmbBFIJF0X9RiYMIgolkwS8Ncbc1fQVyM6d/F1n8zQMRPwNo6wkgn224o8DxpFRbqkAzXUBVq9Vb5f0f/O4eiriU/28AFS5x1LwpxM0ij6Bq9N0SSv4GfmbaJg4LF5mpacS5MogVVpFIhd2YODPCTddf58VRCE1slO4dBweuHGQadYcFDGG68sbR3DHD4qdZAZmPSeGxKz+RLrLxUHhs41AwuEMiDT9rGo3pYBuc2fEyaJo4dkDbTeQrW60qaZsqXpuUr0rqG35vcNBxTQJUK6I16S7qUIAOOwdZ53dUGVVYN2F312cTknJdVRRPz+Ir21WynqnKCWTjw2iFzGfSvYUZOsNSREETFhQDLQ4DkvtRSCGp6HTx3QS+7TEzYYfbVylMtGwG++BQgKJ/rN7aqdaLvknFxm3AXrNu4EOQ2i245JUccCL9/vjPsqPZfq8WE8N42zDTIaaO/hhferw8vJXc1+SbdY027BiQePAfeERAJxbkewxrsdtlHvLJaNkHZdetkzDcBpwcni5/wJxfaw6pGzKbTUnxDDKE=', KEY))