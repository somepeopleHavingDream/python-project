import openpyxl
import os
from Crypto.Cipher import AES
import base64

# xlsx文件路径
# XLSX_PATH = './other_demo/执行结果1.xlsx'
XLSX_PATH = './other_demo/hani_test_charge_channel.xlsx'
# 输出的xlsx文件路径
OUTPUT_PATH = './other_demo/output.xlsx'
# OUTPUT_PATH = './other_demo/output.xlsx'
# 密钥
KEY = 'abc'
# 渠道列
ID_COL = 1
# 渠道名称列
CHANNEL_COL = 2
# attach列
ATTACH_COL = 10

# aes加密
def encrypt(str, key):
    if not key:
        raise ValueError("key不能为空")
    try:
        if not str:
            return None
        # 判断Key是否为16位
        if len(key) != 16:
            return None
        raw = key.encode('utf-8')
        skeySpec = AES.new(raw, AES.MODE_ECB)
        # "算法/模式/补码方式"
        # PKCS#7 padding
        BS = AES.block_size
        padded_str = str.encode('utf-8') + (BS - len(str) % BS) * chr(BS - len(str) % BS).encode('utf-8')
        encrypted = skeySpec.encrypt(padded_str)
        # 此处使用BASE64做转码功能，同时能起到2次加密的作用。
        return base64.b64encode(encrypted).decode('utf-8')
    except Exception as ex:
        return None
    
# 加密attach
def encrypt_attach(file_name, ouput_path):
    # 打开xlsx文件
    wb = openpyxl.load_workbook(file_name)
    # 获取第一个sheet
    sheet = wb.active
    rows = []
    # 遍历每一行数据
    for row in sheet.iter_rows():
        print(f"row length: {len(row)}")
        # 加密每行数据的attach列
        channel = row[ID_COL-1].value
        channel_name = row[CHANNEL_COL-1].value
        attach = row[ATTACH_COL-1].value
        print(channel, channel_name, attach)
        rows.append([channel, channel_name, attach, encrypt(attach, KEY)])
    # 写入到xlsx文件中
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in rows:
        sheet.append(row)
    wb.save(ouput_path)

if __name__ == '__main__':
    # 当前路径
    print(os.getcwd())
    # 读xlsx文件
    encrypt_attach(XLSX_PATH, OUTPUT_PATH)

